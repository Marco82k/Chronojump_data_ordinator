import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils.cell import coordinate_to_tuple
import os
import warnings
import streamlit as st
import io
import traceback
import tempfile
from numbers_parser import Document
import plotly.express as px
import plotly.graph_objects as go
import re

# Ignora avvisi non critici
warnings.filterwarnings("ignore")

# --- CONFIGURAZIONE FILE ---
# Questi rimangono come default o fallback
FILE_SORGENTE_DEFAULT = "Allenamento.xlsx"
FILE_MODELLO_DEFAULT = "excel.xlsx"

# --- CONFIGURAZIONE REGOLA SALTI ---
CONFIG_ANAGRAFICA = [
    # { "etichetta": "Data", "cella": "F2" }, # RIMOSSO: La data viene gestita esternamente
    { "etichetta": "Data di nascita", "cella": "B2" },
    { "etichetta": "Altezza", "cella": "C3" },
    { "etichetta": "Sesso", "cella": "G2" },
    { "etichetta": "Peso", "cella": "C4" },
    { "etichetta": "lunghezza gamba", "cella": "E5" },
    { "etichetta": "altezza dei fianchi durante flessione SJ", "cella": "E6" }
]

REGISTRO_SALTI = [
    # 1. PRIMA SERIE ABK
    {
        "tipo": "ABK", "discriminante": None,
        "outputs": [{"dato": "Altezza", "celle": ["F9", "G9", "H9"]}]
    },
    # 2. PRIMA SERIE CMJ
    {
        "tipo": "CMJ", "discriminante": None,
        "outputs": [{"dato": "Altezza", "celle": ["F10", "G10", "H10"]}]
    },
    # 3. SECONDA SERIE ABK (Se esiste)
    {
        "tipo": "ABK", "discriminante": None,
        "outputs": [{"dato": "Altezza", "celle": ["F12", "G12", "H12"]}]
    },
    # 4. SECONDA SERIE CMJ (Se esiste)
    {
        "tipo": "CMJ", "discriminante": None,
        "outputs": [{"dato": "Altezza", "celle": ["F13", "G13", "H13"]}]
    },
    # 5. SJ
    {
        "tipo": "SJ", "discriminante": None,
        "outputs": [{"dato": "Altezza", "celle": ["S15", "T15", "U15"]}]
    },
    # 6. slCMJ Left
    {
        "tipo": "slCMJleft", "discriminante": None,
        "outputs": [{"dato": "Altezza", "celle": ["F16", "G16", "H16"]}]
    },
    # 7. slCMJ Right
    {
        "tipo": "slCMJright", "discriminante": None,
        "outputs": [{"dato": "Altezza", "celle": ["F15", "G15", "H15"]}]
    },

    # --- DJa (DROP JUMPS) ---
    {
        "tipo": "DJa", "discriminante": ("Caduta", 30),
        "outputs": [{"dato": "Altezza", "celle": ["P4", "Q4", "R4"]}, {"dato": "TC", "celle": ["T4", "U4", "V4"]}]
    },
    {
        "tipo": "DJa", "discriminante": ("Caduta", 45),
        "outputs": [{"dato": "Altezza", "celle": ["P5", "Q5", "R5"]}, {"dato": "TC", "celle": ["T5", "U5", "V5"]}]
    },
    {
        "tipo": "DJa", "discriminante": ("Caduta", 60),
        "outputs": [{"dato": "Altezza", "celle": ["P6", "Q6", "R6"]}, {"dato": "TC", "celle": ["T6", "U6", "V6"]}]
    },
    {
        "tipo": "DJa", "discriminante": ("Caduta", 75),
        "outputs": [{"dato": "Altezza", "celle": ["P7", "Q7", "R7"]}, {"dato": "TC", "celle": ["T7", "U7", "V7"]}]
    },
    {
        "tipo": "DJa", "discriminante": ("Caduta", 90),
        "outputs": [{"dato": "Altezza", "celle": ["P8", "Q8", "R8"]}, {"dato": "TC", "celle": ["T8", "U8", "V8"]}]
    },
    {
        "tipo": "DJa", "discriminante": ("Caduta", 105),
        "outputs": [{"dato": "Altezza", "celle": ["P9", "Q9", "R9"]}, {"dato": "TC", "celle": ["T9", "U9", "V9"]}]
    },

    # --- SJl (SJI - SALTI CON CARICO) ---
    # Logica dinamica: prende le serie SJl in ordine di comparsa nel file
    {
        "tipo": "SJl", "discriminante": None,
        "weight_output": "R16",
        "outputs": [{"dato": "Altezza", "celle": ["S16", "T16", "U16"]}]
    },
    {
        "tipo": "SJl", "discriminante": None,
        "weight_output": "R17",
        "outputs": [{"dato": "Altezza", "celle": ["S17", "T17", "U17"]}]
    },
    {
        "tipo": "SJl", "discriminante": None,
        "weight_output": "R18",
        "outputs": [{"dato": "Altezza", "celle": ["S18", "T18", "U18"]}]
    },
    {
        "tipo": "SJl", "discriminante": None,
        "weight_output": "R19",
        "outputs": [{"dato": "Altezza", "celle": ["S19", "T19", "U19"]}]
    }
]


def custom_round(val, decimals=0):
    """
    Arrotondamento aritmetico:
    - .5 arrotonda sempre per eccesso (es. 2.5 -> 3, 3.5 -> 4)
    - Gestisce sia numeri che stringhe convertibili
    """
    try:
        if val is None or str(val).strip() == "":
            return val
            
        n = float(str(val).replace(',', '.'))
        multiplier = 10 ** decimals
        # Aggiungiamo un epsilon piccolissimo per gestire errori di floating point
        res = int(n * multiplier + 0.5) / multiplier
        
        if decimals == 0:
            return int(res)
        return res
    except:
        return val


def carica_file_universale(uploaded_file):
    """Carica file Excel o CSV da un oggetto file-like di Streamlit o path"""
    if uploaded_file is None:
        return None

    # Se è una stringa (per retrocompatibilità o test locale), lo trattiamo come path
    if isinstance(uploaded_file, str):
        filepath = uploaded_file
        print(f"Lettura file path: {filepath}")
        try:
            return pd.read_excel(filepath, header=None)
        except:
            pass
        for sep in [',', ';', '\t']:
            try:
                df = pd.read_csv(filepath, header=None, sep=sep, encoding='latin1', on_bad_lines='skip', engine='python')
                if df.shape[1] > 1: return df
            except:
                continue
        return None

    # Altrimenti è un buffer di Streamlit
    filename = uploaded_file.name
    print(f"Lettura buffer: {filename}")

    try:
        return pd.read_excel(uploaded_file, header=None)
    except:
        pass

    # Reset del puntatore per tentativi CSV
    uploaded_file.seek(0)

    # Supporto per file Apple .numbers
    if filename.endswith('.numbers'):
        try:
            # numbers-parser richiede un file fisico su disco
            with tempfile.NamedTemporaryFile(delete=False, suffix=".numbers") as tmp:
                tmp.write(uploaded_file.read())
                tmp_path = tmp.name
            
            try:
                doc = Document(tmp_path)
                sheets = doc.sheets
                if sheets:
                    tables = sheets[0].tables
                    if tables:
                        table = tables[0]
                        data = table.rows(values_only=True)
                        df = pd.DataFrame(data)
                        return df
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
                    
        except Exception as e:
            print(f"Errore caricamento .numbers: {e}")

    for sep in [',', ';', '\t']:
        try:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, header=None, sep=sep, encoding='latin1', on_bad_lines='skip', engine='python')
            if df.shape[1] > 1: return df
        except:
            continue
    return None


def trova_valore_cella(df, keywords):
    """(Step 1) Cerca un valore nella griglia usando una o più parole chiave"""
    if isinstance(keywords, str): keywords = [keywords]

    df_str = df.astype(str).apply(lambda x: x.str.lower().str.strip())

    for key in keywords:
        matches = df_str.stack()
        found = matches[matches == key.lower()].index.tolist()

        if not found:
            found = matches[matches.str.contains(key.lower(), regex=False)].index.tolist()

        for (r, c) in found:
            try:
                val = str(df.iloc[r + 1, c]).strip()
                if val not in ["nan", "0", "0.0", "", "None"]:
                    return val
            except:
                continue
    return ""


def raggruppa_salti_per_serie(df_salti):
    """Divide i salti in gruppi contigui."""
    gruppi = []
    if df_salti.empty: return gruppi

    current_chunk = []
    first_row = df_salti.iloc[0]
    last_tipo = first_row['Tipo']
    last_caduta = first_row['Caduta'] if 'Caduta' in df_salti.columns else -1
    last_peso = first_row['Peso Kg'] if 'Peso Kg' in df_salti.columns else -1

    for index, row in df_salti.iterrows():
        curr_tipo = row['Tipo']
        curr_caduta = row['Caduta'] if 'Caduta' in df_salti.columns else -1
        curr_peso = row['Peso Kg'] if 'Peso Kg' in df_salti.columns else -1

        cambio_serie = (curr_tipo != last_tipo) or \
                       (abs(float(curr_caduta) - float(last_caduta)) > 0.1) or \
                       (abs(float(curr_peso) - float(last_peso)) > 0.1)

        if cambio_serie:
            gruppi.append({
                'tipo': last_tipo,
                'caduta': last_caduta,
                'peso': last_peso,
                'data': pd.DataFrame(current_chunk)
            })
            current_chunk = []
            last_tipo = curr_tipo
            last_caduta = curr_caduta
            last_peso = curr_peso

        current_chunk.append(row)

    if current_chunk:
        gruppi.append({
            'tipo': last_tipo,
            'caduta': last_caduta,
            'peso': last_peso,
            'data': pd.DataFrame(current_chunk)
        })

    return gruppi


def elabora_salti_cronologici(df, ws, data_selezionata):
    """Elabora i salti e scrive nel worksheet."""
    st.write("--- ESECUZIONE STEP 2 (ORDINE CRONOLOGICO) ---")

    riga_header = -1
    col_map = {}

    for i, riga in df.iterrows():
        riga_lista = [str(x).strip().lower() for x in riga.tolist()]
        if "tipo" in riga_lista and "altezza" in riga_lista:
            riga_header = i
            for idx, val in enumerate(riga_lista): col_map[val] = idx
            break

    if riga_header == -1:
        st.error("ERRORE: Tabella salti non trovata.")
        return

    df_data = df.iloc[riga_header + 1:].copy()

    def get_col_values(nome_col):
        if nome_col.lower() in col_map: return df_data.iloc[:, col_map[nome_col.lower()]]
        return None

    clean_df = pd.DataFrame()
    clean_df['Tipo'] = get_col_values('Tipo').astype(str).str.strip()

    raw_alt = get_col_values('Altezza')
    clean_df['Altezza'] = pd.to_numeric(raw_alt.astype(str).str.replace(',', '.'), errors='coerce') if raw_alt is not None else 0.0

    raw_tc = get_col_values('TC')
    clean_df['TC'] = pd.to_numeric(raw_tc.astype(str).str.replace(',', '.'), errors='coerce') if raw_tc is not None else 0.0

    raw_caduta = get_col_values('Caduta')
    clean_df['Caduta'] = pd.to_numeric(raw_caduta.astype(str).str.replace(',', '.'), errors='coerce').fillna(-1) if raw_caduta is not None else -1

    raw_peso = get_col_values('Peso Kg')
    if raw_peso is None: raw_peso = get_col_values('Peso')
    clean_df['Peso Kg'] = pd.to_numeric(raw_peso.astype(str).str.replace(',', '.'), errors='coerce').fillna(-1) if raw_peso is not None else -1

    # Filtraggio per data
    raw_data = get_col_values('Data')
    if raw_data is not None:
        clean_df['Data_Originale'] = raw_data.astype(str).str.strip()
        # Convertiamo la data selezionata in stringa per il confronto (formato YYYY-MM-DD)
        # Il formato nel file potrebbe variare, proviamo a normalizzare
        def confronta_date(data_file_str, data_selected):
            try:
                # Prova parsing standard
                d_file = pd.to_datetime(data_file_str).date()
                return d_file == data_selected
            except:
                # Fallback: confronto stringa parziale se fallisce il parsing
                return str(data_selected) in data_file_str

        maschera_data = clean_df['Data_Originale'].apply(lambda x: confronta_date(x, data_selezionata))
        clean_df = clean_df[maschera_data]

    clean_df = clean_df.dropna(subset=['Altezza'])

    gruppi_disponibili = raggruppa_salti_per_serie(clean_df)
    st.info(f"Trovati {len(gruppi_disponibili)} gruppi di salti per la data {data_selezionata}.")

    gruppi_usati = [False] * len(gruppi_disponibili)

    for regola in REGISTRO_SALTI:
        tipo_req = regola['tipo']
        discrim = regola['discriminante']

        gruppo_trovato = None
        idx_trovato = -1

        for i, gruppo in enumerate(gruppi_disponibili):
            if gruppi_usati[i]: continue

            tipo_ok = (gruppo['tipo'].lower() == tipo_req.lower())
            if tipo_req.lower() == "sji":
                tipo_ok = tipo_ok or (gruppo['tipo'].lower() == "sjl")

            if not tipo_ok: continue

            discrim_ok = True
            if discrim:
                nome_d, val_d = discrim
                val_gruppo = gruppo['caduta'] if nome_d == "Caduta" else gruppo['peso']
                if abs(float(val_gruppo) - float(val_d)) > 0.1:
                    discrim_ok = False

            if discrim_ok:
                gruppo_trovato = gruppo['data']
                idx_trovato = i
                break

        if gruppo_trovato is not None:
            gruppi_usati[idx_trovato] = True
            # st.write(f" -> Regola {tipo_req} (Disc: {discrim}): USATO Gruppo {idx_trovato} ({len(gruppo_trovato)} salti)")
            # --- SEZIONE AGGIORNATA ---
            if "weight_output" in regola:
                # Prende il peso dal gruppo corrente (Serie 1, Serie 2, ecc.)
                peso_effettivo = gruppi_disponibili[idx_trovato]['peso']
                # Scrive il peso nella cella R configurata arrotondato a 1 cifra
                try: 
                    peso_effettivo = custom_round(float(str(peso_effettivo).replace(',', '.')), 1)
                except: 
                    pass
                ws[regola["weight_output"]] = peso_effettivo
                if isinstance(peso_effettivo, (int, float)):
                    ws[regola["weight_output"]].number_format = '0.00'
                print(f" -> Scritto peso {peso_effettivo} in {regola['weight_output']}")
            # --------------------------

            # Prende i top 3 valori di Altezza
            # 1. Trova i 3 con altezza maggiore
            top_3_indices = gruppo_trovato.sort_values(by='Altezza', ascending=False).head(3).index 
            # 2. Li ordina per indice (cronologico)
            df_sorted_top = gruppo_trovato.loc[top_3_indices].sort_index()

            for out_conf in regola['outputs']:
                col_dato = out_conf['dato']
                celle = out_conf['celle']
                
                # Determina applicazione arrotondamento
                # Richiesta: NO arrotondamento per DJa (solo dato TC), SI arrotondamento per altri
                no_round_dja = (tipo_req == "DJa" and col_dato == "TC")

                vals_raw = df_sorted_top[col_dato].tolist()
                vals_processed = []
                for v in vals_raw:
                    try:
                        val_float = float(str(v).replace(',', '.'))
                        if no_round_dja:
                            vals_processed.append(val_float)
                        else:
                            # Usa arrotondamento custom
                            vals_processed.append(custom_round(val_float, 1))
                    except:
                        if v is not None and str(v).lower() != "nan":
                            vals_processed.append(v)
                
                # NUOVA LOGICA: Gestione 1 o 2 dati per riempire 3 caselle
                if len(vals_processed) == 1:
                    # Caso 1 dato -> Lo ripetiamo per 3 volte
                    val = vals_processed[0]
                    vals_processed = [val, val, val]
                elif len(vals_processed) == 2:
                    # Caso 2 dati -> Facciamo la media tra i due per il terzo dato
                    try:
                        v1 = float(vals_processed[0])
                        v2 = float(vals_processed[1])
                        media = (v1 + v2) / 2
                        if not no_round_dja:
                            media = custom_round(media, 1)
                        vals_processed.append(media)
                    except:
                        pass
                
                # Ordinamento RIMOSSO per mantenere cronologia
                # vals_processed.sort()

                for k, cella in enumerate(celle):
                    val = vals_processed[k] if k < len(vals_processed) else ""
                    ws[cella] = val
                    if isinstance(val, (int, float)):
                        if no_round_dja:
                            ws[cella].number_format = '0.000'
                        else:
                            ws[cella].number_format = '0.00'
        else:
            # st.warning(f" -> Regola {tipo_req} (Disc: {discrim}): NESSUN GRUPPO TROVATO (Lascio bianco)")
            for out_conf in regola['outputs']:
                for cella in out_conf['celle']:
                    ws[cella] = ""


def elabora_salti_rj(df, ws, data_selezionata):
    """
    Elabora i salti reattivi (RJ) con LOGICA RIGOROSA A COORDINATE RELATIVE:
    1. Cerca riga con 'RJ'/'RJ(unlimited)' nella colonna 'Tipo di salto'.
    2. Riga+1: Verifica intestazioni colonne -> Col B (TC), Col D (Altezza), Col E (RSI).
    3. Riga+4: Verifica ancoraggio 'SD' in Col A.
    4. Riga+5: Inizio dati numerici. Legge finché Col A contiene numeri.
    5. Calcola Top 5 e scrive risultati.
    """
    st.write("--- ESECUZIONE STEP 3 (RJ: COORDINATE RIGIDE) ---")
    
    from openpyxl.styles import Alignment

    # 1. Trova Colonna "Tipo di salto"
    idx_tipo = -1
    idx_data = -1
    
    # Cerchiamo l'intestazione generale per capire dov'è la colonna Tipo
    for i, riga in df.iterrows():
        riga_lista = [str(x).strip().lower() for x in riga.tolist()]
        if "tipo di salto" in riga_lista:
            # Trovata riga header principale
            for idx, val in enumerate(riga_lista):
                if "tipo di salto" in val: idx_tipo = idx
                if "data" in val: idx_data = idx
            break
    
    if idx_tipo == -1:
        st.warning("⚠️ Colonna 'Tipo di salto' non identificata nel file.")
        return

    # Helper functions
    def to_float(val):
        try:
            return float(str(val).replace(',', '.').strip())
        except:
            return None

    def check_date_match(val_cella, target):
        try:
            s = str(val_cella).strip()
            if pd.to_datetime(s, dayfirst=True, errors='coerce').date() == target: return True
            if str(target) in s: return True
        except:
            pass
        return False

    def is_number(val):
        try:
            float(str(val).replace(',', '.').strip())
            return True
        except:
            return False

    n_rows = len(df)
    sessions_found = []
    
    # Scansione per trovare righe RJ
    i = 0
    while i < n_rows:
        row = df.iloc[i]
        
        # Check RJ
        is_rj = False
        try:
            val_tipo = str(row.iloc[idx_tipo]).strip().lower()
            if "rj" in val_tipo:
                # Check Date
                val_data = row.iloc[idx_data] if idx_data != -1 else None
                if idx_data != -1 and check_date_match(val_data, data_selezionata):
                    is_rj = True
                elif idx_data == -1:
                    # Se non abbiamo colonna data, controlliamo nella riga? 
                    # Assumiamo che ci sia la colonna data se abbiamo trovato 'Tipo di salto'
                    pass
        except: pass
        
        if is_rj:
            st.info(f"📍 Trovato potenziale RJ a riga {i+1}. Verifico struttura...")
            
            # --- VERIFICA COORDINATE RELATIVE ---
            
            # 1. Riga + 1: Colonne B (TC), D (Altezza), E (RSI)
            # Indici (0-based): B=1, D=3, E=4
            r_sigle = i + 1
            if r_sigle >= n_rows: break
            
            row_sigle = df.iloc[r_sigle]
            
            # Controllo "blando" sulle sigle per confermare che siamo nel posto giusto
            try:
                sigla_b = str(row_sigle.iloc[1]).strip().lower() # TC
                sigla_d = str(row_sigle.iloc[3]).strip().lower() # Altezza
                sigla_e = str(row_sigle.iloc[4]).strip().lower() # RSI
                
                # Keywords
                ok_b = any(x in sigla_b for x in ['tc', 'contact', 'time'])
                ok_d = any(x in sigla_d for x in ['altezza', 'height', 'h '])
                ok_e = any(x in sigla_e for x in ['rsi', 'reactive', 'index'])
                
                if not (ok_b and ok_d): 
                    st.warning(f"   Struttura colonne non corrispondente a riga {r_sigle+1} (D={sigla_d}, B={sigla_b})")
                    i += 1
                    continue
            except:
                i += 1
                continue

            # 2. Riga + 4 (partendo da 'i+1' scendo di 3 -> i+1+3 = i+4): Ancoraggio 'SD' in Col A (0)
            r_sd = i + 4
            if r_sd >= n_rows: break
            
            row_sd = df.iloc[r_sd]
            try:
                sigla_sd = str(row_sd.iloc[0]).strip().lower()
                if "sd" not in sigla_sd and "jump" not in sigla_sd:
                     st.warning(f"   Manca 'SD' in colonna A alla riga {r_sd+1}")
                     i += 1
                     continue
            except:
                i += 1
                continue
                
            # 3. Riga + 5: Inizio Dati
            r_data = i + 5
            collected_jumps = []
            
            k = r_data
            while k < n_rows:
                d_row = df.iloc[k]
                
                # Check Colonna A (Indice Salto)
                val_a = d_row.iloc[0]
                if not is_number(val_a):
                     # Fine dati
                     break
                
                # Estrazione Valori (B=1 TC, D=3 H, E=4 RSI)
                try:
                    val_tc = to_float(d_row.iloc[1])
                    val_h  = to_float(d_row.iloc[3])
                    val_rsi= to_float(d_row.iloc[4])
                    
                    if val_h is not None and val_h > 0:
                        collected_jumps.append({
                            'h': val_h,
                            'tc': val_tc if val_tc else 0.0,
                            'rsi': val_rsi if val_rsi else 0.0
                        })
                except:
                    pass
                
                k += 1
            
            # --- CALCOLO STATISTICHE SESSIONE ---
            if collected_jumps:
                df_sess = pd.DataFrame(collected_jumps)
                
                # 1. Filtro Qualità: Scarta H <= 0 o NaN
                df_sess = df_sess[df_sess['h'] > 0].copy()
                
                if not df_sess.empty:
                    # 2. Seleziona Top 5 per Altezza
                    if len(df_sess) > 5:
                        top_5 = df_sess.sort_values(by='h', ascending=False).head(5)
                    else:
                        top_5 = df_sess
                    
                    # 3. Calcolo Medie Robuste (escludendo 0 nel conteggio per TC e RSI)
                    # Helper per media non-zero
                    def mean_exclude_zeros(series):
                        valid_vals = series[series > 0]
                        if valid_vals.empty: return 0.0
                        return valid_vals.mean()

                    avg_h = top_5['h'].mean()
                    avg_tc = mean_exclude_zeros(top_5['tc'])
                    avg_rsi = mean_exclude_zeros(top_5['rsi'])
                    
                    stats = {
                        'avg_h': avg_h,
                        'avg_tc': avg_tc,
                        'avg_rsi': avg_rsi,
                        'n_salti': len(df_sess), # Salti validi totali
                        'start_row': i
                    }
                    sessions_found.append(stats)
                    st.success(f"   ✅ Sessione valida estratta: {len(df_sess)} salti validi.")
                
            # Saltiamo k righe
            i = k
            continue
            
        i += 1
        
    # --- SELEZIONE MIGLIORE E SCRITTURA ---
    if not sessions_found:
        st.warning(f"Nessuna sessione RJ valida trovata per la data {data_selezionata}.")
        ws["F19"] = ""; ws["H19"] = ""; ws["I19"] = ""
        return

    # Migliore per Avg H
    best = max(sessions_found, key=lambda x: x['avg_h'])
    
    st.markdown(f"**🏆 Sessione Vincente (Riga {best['start_row']+1}):** Avg H {best['avg_h']:.2f}")

    # Scrittura
    ws["F19"] = custom_round(best['avg_h'], 2)
    ws["F19"].number_format = '0.00'
    
    if best['avg_tc'] > 0:
        ws["H19"] = custom_round(best['avg_tc'], 3)
        ws["H19"].number_format = '0.000'
    else:
        ws["H19"] = ""
    
    if best['avg_rsi'] > 0:
        ws["I19"] = custom_round(best['avg_rsi'], 3)
        ws["I19"].number_format = '0.000'
    else:
        ws["I19"] = ""


def elabora_step1_anagrafica(df, ws, data_selezionata):
    """Elabora l'anagrafica leggendo ESCLUSIVAMENTE la riga sotto 'ID'"""
    st.write("--- ESECUZIONE STEP 1 (ANAGRAFICA RIGIDA) ---")

    # 1. Data Test (F2)
    ws["F2"] = data_selezionata.strftime("%d/%m/%Y")

    # 2. Trova la riga dell'intestazione (dove c'è scritto ID, Nome, Altezza...)
    riga_header = -1
    col_map = {}
    for i, riga in df.iterrows():
        riga_check = [str(x).strip().lower() for x in riga.tolist()]
        # Cerchiamo la riga che contiene i metadati dell'atleta
        if "id" in riga_check and "nome" in riga_check:
            riga_header = i
            for idx, col_name in enumerate(riga_check):
                col_map[col_name] = idx
            break

    if riga_header == -1:
        st.error("ERRORE: Riga 'ID' non trovata. Impossibile leggere l'altezza corretta.")
        return

    # La riga dell'atleta è quella immediatamente sotto l'header ID
    riga_atleta = df.iloc[riga_header + 1]

    # FUNZIONE LOCALE: Prende il dato SOLO dalla riga dell'atleta
    def prendi_solo_da_riga_id(nome_colonna):
        n = nome_colonna.lower().strip()
        if n in col_map:
            idx = col_map[n]
            val = str(riga_atleta.iloc[idx]).strip()
            # Il valore se 0, nan, none o vuoto deve essere "" (BIANCO)
            if val.lower() in ["nan", "none", "", "0", "0.0"]:
                return ""
            return val
        return ""

    # 3. Nome e Cognome (C1, E1)
    full_name = prendi_solo_da_riga_id("nome")
    if not full_name: full_name = prendi_solo_da_riga_id("nome persona")
    
    if full_name:
        parti = full_name.split(" ", 1)
        ws["C1"] = parti[0].strip().upper() if len(parti) > 0 else ""
        ws["E1"] = parti[1].strip().upper() if len(parti) > 1 else ""
        st.info(f" -> Splittato nome: {ws['C1'].value} (C1), {ws['E1'].value} (E1)")

    # 4. Ciclo sui campi (Altezza, Peso, ecc.) usando SOLO la riga ID
    for item in CONFIG_ANAGRAFICA:
        etichetta = item['etichetta']
        cella = item['cella']
        
        valore = prendi_solo_da_riga_id(etichetta)
        
        # Fallback per il Peso
        if not valore and etichetta.lower() == "peso":
            valore = prendi_solo_da_riga_id("peso kg")

        # Conversione Sesso
        if etichetta.lower() == "sesso":
            if valore.upper() == "M": valore = "UOMO"
            elif valore.upper() == "F": valore = "DONNA"
            
        
        # Arrotondamento anagrafica: RIMUOVERE LA VIRGOLA (Interi)
        # Richiesta: "levare i numeri con la virgola dai dati anagrafici" -> custom_round(val, 0)
        is_number = False
        try:
            val_num = float(str(valore).replace(',', '.'))
            valore = custom_round(val_num, 0) # Arrotondamento all'intero
            is_number = True
        except:
            pass

        ws[cella] = valore
        if is_number:
            ws[cella].number_format = '0' # Formato intero senza decimali
        
        st.info(f" -> {etichetta}: {valore} (scritto in {cella})")


def main():
    st.set_page_config(page_title="Athletic Data Excel Sync 📈", page_icon="🚀", layout="wide")
    
    # --- Sidebar Navigation ---
    st.sidebar.title("Menu Navigazione")
    pagina = st.sidebar.radio("Vai a:", ["Athletic Data", "Report"])

    if pagina == "Athletic Data":
        # --- Pagina Principale (Codice Esistente) ---
        st.title("🚀 Athletic Data Excel Sync 📈")
        st.markdown("Carica il file dati e il modello Excel per generare il report finale.")

        # 1. Widget Caricamento File Sorgente
        uploaded_file_sorgente = st.file_uploader("Carica file 'Allenamento' (Excel, CSV o Numbers)", type=['xlsx', 'csv', 'numbers'])

        # 2. Widget Caricamento File Modello
        uploaded_file_modello = st.file_uploader("Carica file 'Modello' (Excel)", type=['xlsx'])

        # Opzione per usare modello locale se non caricato
        path_modello_locale = FILE_MODELLO_DEFAULT
        uso_modello_locale = False

        if not uploaded_file_modello:
            if os.path.exists(path_modello_locale):
                st.info(f"Nessun modello caricato. Verrà usato '{path_modello_locale}' se presente nella cartella.")
                uso_modello_locale = True
            else:
                st.warning(f"Attenzione: Modello '{path_modello_locale}' non trovato in locale. Caricane uno.")

        # 3. Data dei Dati
        col1, col2 = st.columns(2)
        with col1:
            data_test = st.date_input("📅 Seleziona Data Test", value=pd.Timestamp.now().date(), format="DD/MM/YYYY")
        
        # 4. Nome Output
        with col2:
            output_name = st.text_input("Nome file di output", value="Risultati_Atleta.xlsx")
            if not output_name.endswith(".xlsx"):
                output_name += ".xlsx"

        # 5. Pulsante elaborazione
        if st.button("Avvia Elaborazione", type="primary"):
            if not uploaded_file_sorgente:
                st.error("Per favore carica il file sorgente 'Allenamento'.")
                return

            # Determinazione del modello da usare
            modello_da_usare = None
            if uploaded_file_modello:
                modello_da_usare = uploaded_file_modello
            elif uso_modello_locale:
                modello_da_usare = path_modello_locale
            else:
                st.error("Manca il file Modello! Caricalo.")
                return

            with st.spinner("Elaborazione in corso..."):
                try:
                    # A. Caricamento Dati
                    df = carica_file_universale(uploaded_file_sorgente)
                    if df is None:
                        st.error("Errore lettura file sorgente. Verifica il formato.")
                        return

                    # B. Caricamento Modello e Processamento
                    wb = load_workbook(modello_da_usare)
                    # Seleziona sempre il primo foglio disponibile
                    if len(wb.worksheets) > 0:
                        ws = wb.worksheets[0]
                        st.info(f"Foglio selezionato automaticamente: {ws.title}")
                    else:
                        st.error("Il file modello non contiene fogli di lavoro.")
                        return

                    # C. Esecuzione Step
                    elabora_step1_anagrafica(df, ws, data_test)
                    elabora_salti_cronologici(df, ws, data_test)
                    elabora_salti_rj(df, ws, data_test)

                    # D. Salvataggio in memoria (BytesIO)
                    buffer = io.BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)

                    st.success("Elaborazione Completata con Successo! ✅")

                    # E. Bottone Download
                    st.download_button(
                        label="📥 Scarica File Elaborato",
                        data=buffer,
                        file_name=output_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                except Exception as e:
                    st.error(f"Errore durante l'elaborazione: {e}")
                    st.text(traceback.format_exc())

    elif pagina == "Report":
        
        st.title("📊 Report Comparativo PRE/POST")
        st.markdown("Carica i file PRE e POST per generare il confronto automatico.")

        # 1. Uploaders
        col_up1, col_up2 = st.columns(2)
        with col_up1:
            file_pre = st.file_uploader("📂 Carica File PRE (Start)", type=['xlsx', 'csv', 'numbers'])
        with col_up2:
            file_post = st.file_uploader("📂 Carica File POST (End)", type=['xlsx', 'csv', 'numbers'])
        
        file_template = st.file_uploader("📂 Carica Template Report (Opzionale - Default: report.xlsx)", type=['xlsx'])

        # 2. Bottone Generazione
        if st.button("Genera Report Comparativo", type="primary"):
            if not file_pre or not file_post:
                st.error("⚠️ Per favore carica entrambi i file PRE e POST.")
            else:
                with st.spinner("⏳ Generazione Report in corso..."):
                    try:
                        # A. Caricamento DataFrames
                        df_pre = carica_file_universale(file_pre)
                        df_post = carica_file_universale(file_post)
                        
                        if df_pre is None or df_post is None:
                            st.error("Errore nella lettura dei file. Verifica il formato.")
                            st.stop()

                        # B. Caricamento Template
                        wb_report = None
                        if file_template:
                            wb_report = load_workbook(file_template)
                        elif os.path.exists("report.xlsx"):
                            wb_report = load_workbook("report.xlsx")
                        else:
                            st.warning("⚠️ Template 'report.xlsx' non trovato. Creazione nuovo file vuoto.")
                            from openpyxl import Workbook
                            wb_report = Workbook()
                        
                        ws_report = wb_report.active

                        # C. LOGICA ESTRAZIONE & SCRITTURA
                        import re # Import locale per sicurezza
                        
                        MAPPING_CONFIG = [
                            {"label": "PESO", "row_report": 2, "cell_source": "C4"},
                            {"label": "COSCIA DX", "row_report": 3, "cell_source": "G5"},
                            {"label": "COSCIA SX", "row_report": 4, "cell_source": "H5"},
                            {"label": "CMJ OPEN SQUAT NA [ABK]", "row_report": 5, "cell_source": "J9"},
                            {"label": "CMJ HALF SQUAT NA [CMJ]", "row_report": 6, "cell_source": "J10"},
                            {"label": "CMJ OPEN SQUAT BL [ABK]", "row_report": 7, "cell_source": "J12"},
                            {"label": "CMJ HALF SQUAT BL [CMJ]", "row_report": 8, "cell_source": "J13"},
                            {"label": "SL CMJ DX BL", "row_report": 9, "cell_source": "I15"},
                            {"label": "SL CMJ SX BL", "row_report": 10, "cell_source": "I16"},
                            {"label": "RJ [Unlimited]", "row_report": 11, "cell_source": "F19"},
                            {"label": "Vertec - SAM", "row_report": 12, "cell_source": "E26"},
                            {"label": "Vertec - SDA", "row_report": 13, "cell_source": "E27"},
                            {"label": "DJa 30cm", "row_report": 14, "cell_source": "S4"},
                            {"label": "DJa 45cm", "row_report": 15, "cell_source": "S5"},
                            {"label": "DJa 60cm", "row_report": 16, "cell_source": "S6"},
                            {"label": "DJa 75cm", "row_report": 17, "cell_source": "S7"},
                            {"label": "DJa 90cm", "row_report": 18, "cell_source": "S8"},
                            {"label": "DJa 105cm", "row_report": 19, "cell_source": "S9"},
                            {"label": "SJ", "row_report": 20, "cell_source": "W15"},
                            {"label": "SJi 25%", "row_report": 21, "cell_source": "W16"},
                            {"label": "SJi 50%", "row_report": 22, "cell_source": "W17"},
                            {"label": "SJi 75%", "row_report": 23, "cell_source": "W18"},
                            {"label": "SJi 100%", "row_report": 24, "cell_source": "W19"},
                            {"label": "1RM", "row_report": 25, "cell_source": "U28"},
                        ]

                        # --- 1. Funzione di Pulizia Avanzata ---
                        def clean_numeric_value(val):
                            """
                            Pulisce il valore da testo (kg, cm, etc), converte virgola in punto
                            e restituisce float. Se fallisce restituisce 0.0.
                            """
                            if val is None: return 0.0
                            s = str(val).strip()
                            if s == "": return 0.0
                            
                            # Rimuove tutto tranne numeri, punto, virgola, segno meno
                            s_clean = re.sub(r'[^\d.,\-]', '', s)
                            if not s_clean: return 0.0
                            
                            # Sostituisce virgola con punto
                            s_clean = s_clean.replace(',', '.')
                            
                            try:
                                return float(s_clean)
                            except:
                                return 0.0

                        # --- 2. Helper Caricamento Excel Robusto ---
                        def load_excel_robust(file_upl, nome_log):
                            """
                            Carica WB DataOnly (per valori) e WB Formule (per check).
                            Cerca foglio 'ATLETA' (case insensitive).
                            Restituisce (ws_val, ws_form, error_msg)
                            """
                            if not file_upl.name.lower().endswith('.xlsx'):
                                return None, None, "Not XLSX"

                            try:
                                # WB Valori (Data Only = True)
                                file_upl.seek(0)
                                wb_val = load_workbook(file_upl, data_only=True)
                                
                                # WB Formule (Data Only = False)
                                file_upl.seek(0)
                                wb_form = load_workbook(file_upl, data_only=False)

                                # Ricerca Foglio
                                sheet_name = None
                                for s in wb_val.sheetnames:
                                    if "atleta" in s.lower().strip():
                                        sheet_name = s
                                        break
                                
                                if not sheet_name:
                                    sheet_name = wb_val.sheetnames[0]
                                    st.warning(f"⚠️ Nel file {nome_log} non ho trovato il foglio 'ATLETA'. Uso il primo foglio: '{sheet_name}'")
                                else:
                                    st.info(f"✅ File {nome_log}: trovato foglio target '{sheet_name}'")

                                return wb_val[sheet_name], wb_form[sheet_name], None

                            except Exception as e:
                                return None, None, str(e)

                        # --- 3. Caricamento Workbooks (Una volta sola) ---
                        ws_pre_val, ws_pre_form, err_pre = load_excel_robust(file_pre, "PRE")
                        ws_post_val, ws_post_form, err_post = load_excel_robust(file_post, "POST")
                        
                        # --- ESTRAZIONE NOME ATLETA PER FILE ---
                        nome_atleta = "Atleta_Anonimo"
                        try:
                            # Cerchiamo prima nel POST, poi nel PRE
                            target_ws = ws_post_val if ws_post_val else ws_pre_val
                            if target_ws:
                                cognome = str(target_ws["C1"].value or "").strip()
                                nome = str(target_ws["E1"].value or "").strip()
                                
                                if cognome or nome:
                                    # Unisci e pulisci spazi/caratteri strani
                                    full = f"{cognome}_{nome}".strip('_')
                                    # Rimuovi char non validi per filename
                                    nome_atleta = re.sub(r'[^\w\-]', '', full.replace(' ', '_'))
                        except Exception as e_name:
                            print(f"Errore estrazione nome: {e_name}")

                        # Se non sono XLSX, avremo df_pre e df_post (già caricati sopra con carica_file_universale)
                        # Ma per coerenza usiamo logica dedicata
                        
                        # Font Colors
                        RED_FONT = Font(color="FF0000", bold=True)
                        GREEN_FONT = Font(color="00B050", bold=True) # Verde Excel standard
                        
                        # Intestazioni Colonne Report
                        ws_report["B1"] = "PRIMA"
                        ws_report["C1"] = "DOPO"
                        ws_report["D1"] = "RISULTATI"
                        ws_report["E1"] = "RISULTATI %"

                        # --- LISTA PER ANTEPRIMA ---
                        preview_data = []

                        # --- 4. Iterazione Mapping ---
                        for mappa in MAPPING_CONFIG:
                            r_idx = mappa["row_report"]
                            label = mappa.get("label", "")
                            coord = mappa["cell_source"]
                            
                            # Etichetta Report
                            cell_label = ws_report.cell(row=r_idx, column=1)
                            if not cell_label.value:
                                cell_label.value = label

                            # --- Estrazione PRE ---
                            raw_pre = None
                            if ws_pre_val: # Uso Excel OpenPyXL
                                cell_v = ws_pre_val[coord]
                                raw_pre = cell_v.value
                            else: # Uso DataFrame
                                try:
                                    r, c = coordinate_to_tuple(coord)
                                    raw_pre = df_pre.iloc[r-1, c-1] if df_pre is not None else 0
                                except: raw_pre = 0

                            val_pre = clean_numeric_value(raw_pre)

                            # --- Estrazione POST ---
                            raw_post = None
                            if ws_post_val: # Uso Excel OpenPyXL
                                cell_v = ws_post_val[coord]
                                raw_post = cell_v.value
                            else:
                                try:
                                    r, c = coordinate_to_tuple(coord)
                                    raw_post = df_post.iloc[r-1, c-1] if df_post is not None else 0
                                except: raw_post = 0

                            val_post = clean_numeric_value(raw_post)
                            
                            # 1. Scrittura Colonna B (PRIMA) e C (DOPO)
                            cell_prima = ws_report.cell(row=r_idx, column=2)
                            cell_dopo = ws_report.cell(row=r_idx, column=3)
                            
                            cell_prima.value = val_pre
                            cell_dopo.value = val_post
                            
                            cell_prima.number_format = '0.00'
                            cell_dopo.number_format = '0.00'

                            # 2. Calcolo Differenza (Colonna D = 4)
                            diff = val_post - val_pre
                            cell_diff = ws_report.cell(row=r_idx, column=4)
                            cell_diff.value = diff
                            cell_diff.number_format = '0.00'
                            
                            if diff < 0:
                                cell_diff.font = RED_FONT
                            else:
                                cell_diff.font = GREEN_FONT

                            # 3. Calcolo % (Colonna E = 5)
                            # (Post - Pre) / Pre
                            cell_perc = ws_report.cell(row=r_idx, column=5)
                            perc_val = 0.0
                            if val_pre != 0:
                                perc = (diff / val_pre) # Decimale
                                perc_val = perc
                                cell_perc.value = perc
                                cell_perc.number_format = '0.00%'
                                
                                if perc < 0:
                                    cell_perc.font = RED_FONT
                                else:
                                    cell_perc.font = GREEN_FONT
                            else:
                                cell_perc.value = ""
                            
                            # Aggiungi riga ai dati per anteprima
                            preview_data.append({
                                "Test": label,
                                "PRIMA": val_pre,
                                "DOPO": val_post,
                                "Diff": diff,
                                "Diff %": f"{perc_val:.2%}" if val_pre != 0 else "",
                                "PercRaw": perc_val * 100 # Salviamo come numero (es. 10.5) per i grafici
                            })
                        
                        # --- ANTEPRIMA ---
                        if preview_data:
                            st.write("### 📂 Anteprima Report Generato")
                            df_dashboard = pd.DataFrame(preview_data)
                            
                            # Mostra Tabella (senza colonna PercRaw che è tecnica)
                            st.dataframe(df_dashboard.drop(columns=["PercRaw"]), width="stretch")


                        # D. Salvataggio
                        buffer = io.BytesIO()
                        wb_report.save(buffer)
                        buffer.seek(0)
                        
                        st.success("Report Generato con Successo!")
                        st.download_button(
                            label="📥 Scarica Report Comparativo",
                            data=buffer,
                            file_name=f"Report_Analisi_{nome_atleta}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                    except Exception as e:
                        st.error(f"Errore durante l'elaborazione del report: {e}")
                        st.write(traceback.format_exc())

if __name__ == "__main__":
    main()
